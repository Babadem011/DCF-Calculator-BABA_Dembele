#!/usr/bin/env python3
"""
DCF Calculator Universal - Version Compatible Tous Environnements
Calculateur DCF professionnel qui s'adapte automatiquement aux modules disponibles

Version: 2.0 Universal - Fonctionne avec Python seul ou avec modules avancés
Fonctionnalités: DCF Multi-Phases | Monte Carlo | WACC Détaillé | Sensibilité 2D | Auto-détection
"""

import math
import json
import random
import statistics
import sys
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Union

# ==================== DÉTECTION AUTOMATIQUE DES MODULES ====================

# Détection des modules optionnels
HAS_MATPLOTLIB = False
HAS_PANDAS = False
HAS_NUMPY = False
HAS_TABULATE = False
HAS_OPENPYXL = False

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    HAS_MATPLOTLIB = True
    print("✅ Matplotlib détecté - Graphiques avancés disponibles")
except ImportError:
    print("⚠️ Matplotlib non disponible - Mode texte uniquement")

try:
    import pandas as pd
    HAS_PANDAS = True
    print("✅ Pandas détecté - Traitement de données avancé")
except ImportError:
    print("⚠️ Pandas non disponible - Utilisation de structures Python natives")

try:
    import numpy as np
    HAS_NUMPY = True
    print("✅ NumPy détecté - Calculs numériques optimisés")
except ImportError:
    print("⚠️ NumPy non disponible - Utilisation des fonctions Python natives")
    # Fallback pour numpy
    class FakeNumPy:
        @staticmethod
        def arange(start, stop, step):
            result = []
            current = start
            while current < stop:
                result.append(current)
                current += step
            return result
        
        @staticmethod
        def percentile(data, percentile):
            sorted_data = sorted(data)
            index = (percentile / 100) * (len(sorted_data) - 1)
            if index.is_integer():
                return sorted_data[int(index)]
            else:
                lower = sorted_data[int(index)]
                upper = sorted_data[int(index) + 1]
                return lower + (upper - lower) * (index - int(index))
    
    np = FakeNumPy()

try:
    from tabulate import tabulate
    HAS_TABULATE = True
    print("✅ Tabulate détecté - Tableaux formatés")
except ImportError:
    print("⚠️ Tabulate non disponible - Utilisation de tableaux simples")
    # Fallback pour tabulate
    def tabulate(data, headers=None, tablefmt="grid"):
        if not headers:
            headers = [f"Col{i+1}" for i in range(len(data[0]) if data else 0)]
        
        # Calcul des largeurs de colonnes
        all_rows = [headers] + [[str(cell) for cell in row] for row in data]
        col_widths = [max(len(str(row[i])) for row in all_rows) for i in range(len(headers))]
        
        # Ligne de séparation
        separator = "+" + "+".join("-" * (width + 2) for width in col_widths) + "+"
        
        result = [separator]
        
        # En-têtes
        header_row = "|" + "|".join(f" {headers[i]:<{col_widths[i]}} " for i in range(len(headers))) + "|"
        result.append(header_row)
        result.append(separator)
        
        # Données
        for row in data:
            data_row = "|" + "|".join(f" {str(row[i]):<{col_widths[i]}} " for i in range(len(row))) + "|"
            result.append(data_row)
        
        result.append(separator)
        return "\n".join(result)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
    print("✅ OpenPyXL détecté - Export Excel disponible")
except ImportError:
    print("⚠️ OpenPyXL non disponible - Export CSV disponible")

print(f"\n🚀 Mode de fonctionnement: {'COMPLET' if all([HAS_MATPLOTLIB, HAS_PANDAS, HAS_NUMPY, HAS_TABULATE]) else 'ADAPTATIF'}")
print("-" * 60)

# ==================== CONFIGURATION AVANCÉE ====================

INDUSTRY_BENCHMARKS = {
    "Technology": {
        "ebitda_margin": (15, 35), "revenue_growth": (5, 25), "wacc": (8, 15),
        "terminal_growth": (1.5, 3.5), "ev_revenue": (3, 12), "ev_ebitda": (8, 25),
        "beta": (1.2, 2.0), "debt_equity": (0.1, 0.4), "capex_percent": (3, 8),
        "nwc_percent": (5, 15), "tax_rate": (20, 28), "roic": (15, 35)
    },
    "Healthcare": {
        "ebitda_margin": (12, 28), "revenue_growth": (3, 15), "wacc": (7, 12),
        "terminal_growth": (2, 4), "ev_revenue": (2, 8), "ev_ebitda": (6, 18),
        "beta": (0.8, 1.3), "debt_equity": (0.2, 0.5), "capex_percent": (4, 10),
        "nwc_percent": (8, 20), "tax_rate": (22, 30), "roic": (12, 25)
    },
    "Manufacturing": {
        "ebitda_margin": (8, 18), "revenue_growth": (2, 12), "wacc": (6, 11),
        "terminal_growth": (1.5, 3), "ev_revenue": (1, 4), "ev_ebitda": (4, 12),
        "beta": (1.0, 1.5), "debt_equity": (0.3, 0.7), "capex_percent": (6, 12),
        "nwc_percent": (10, 25), "tax_rate": (25, 32), "roic": (8, 18)
    },
    "Retail": {
        "ebitda_margin": (5, 15), "revenue_growth": (1, 8), "wacc": (7, 13),
        "terminal_growth": (1, 3), "ev_revenue": (0.5, 2.5), "ev_ebitda": (3, 10),
        "beta": (1.1, 1.6), "debt_equity": (0.2, 0.6), "capex_percent": (3, 7),
        "nwc_percent": (5, 15), "tax_rate": (24, 30), "roic": (10, 20)
    },
    "Energy": {
        "ebitda_margin": (20, 40), "revenue_growth": (-5, 15), "wacc": (8, 14),
        "terminal_growth": (1, 3), "ev_revenue": (1, 3), "ev_ebitda": (3, 8),
        "beta": (1.3, 2.2), "debt_equity": (0.2, 0.8), "capex_percent": (15, 25),
        "nwc_percent": (5, 15), "tax_rate": (25, 35), "roic": (5, 15)
    },
    "Financial": {
        "ebitda_margin": (25, 45), "revenue_growth": (2, 12), "wacc": (8, 12),
        "terminal_growth": (2, 4), "ev_revenue": (2, 6), "ev_ebitda": (6, 15),
        "beta": (1.0, 1.8), "debt_equity": (0.1, 0.3), "capex_percent": (2, 5),
        "nwc_percent": (0, 5), "tax_rate": (20, 30), "roic": (8, 18)
    },
    "General": {
        "ebitda_margin": (10, 25), "revenue_growth": (2, 15), "wacc": (7, 13),
        "terminal_growth": (1.5, 3.5), "ev_revenue": (1, 6), "ev_ebitda": (4, 15),
        "beta": (1.0, 1.5), "debt_equity": (0.2, 0.5), "capex_percent": (4, 8),
        "nwc_percent": (8, 18), "tax_rate": (24, 30), "roic": (10, 20)
    }
}

MARKET_DATA = {
    "risk_free_rate": 3.5,  # Taux sans risque (OAT 10 ans)
    "market_risk_premium": 6.0,  # Prime de risque de marché
    "gdp_growth": 2.0,  # Croissance PIB long terme
    "inflation": 2.0  # Inflation cible BCE
}

# ==================== CLASSES UTILITAIRES ====================

class SimpleTable:
    """Gestionnaire de tableaux simple pour compatibilité universelle"""
    
    @staticmethod
    def create_table(data, headers, title=""):
        if HAS_TABULATE:
            return tabulate(data, headers=headers, tablefmt="grid")
        else:
            return tabulate(data, headers=headers, tablefmt="grid")

class UniversalVisualizer:
    """Gestionnaire de visualisations adaptatif"""
    
    def __init__(self):
        self.can_plot = HAS_MATPLOTLIB
    
    def create_visualizations(self, dcf_calculator):
        """Crée des visualisations si matplotlib est disponible"""
        if not self.can_plot:
            print("\n📊 VISUALISATIONS NON DISPONIBLES")
            print("Pour activer les graphiques, installez: pip install matplotlib seaborn")
            print("Les analyses restent complètes en mode texte.")
            return False
        
        try:
            return self._create_advanced_plots(dcf_calculator)
        except Exception as e:
            print(f"⚠️ Erreur lors de la création des graphiques: {e}")
            return False
    
    def _create_advanced_plots(self, calc):
        """Génère les graphiques avancés"""
        plt.style.use('default')
        fig = plt.figure(figsize=(20, 16))
        
        # 1. Évolution des métriques
        ax1 = plt.subplot(3, 4, 1)
        years = [proj['year'] for proj in calc.projections]
        revenues = [proj['revenue'] for proj in calc.projections]
        fcfs = [proj['fcf'] for proj in calc.projections]
        
        ax1.plot(years, revenues, 'b-o', linewidth=2, label='CA (M€)')
        ax1_twin = ax1.twinx()
        ax1_twin.plot(years, fcfs, 'g-s', linewidth=2, label='FCF (M€)')
        ax1.set_title('Évolution CA vs FCF', fontweight='bold')
        ax1.legend(loc='upper left')
        ax1_twin.legend(loc='upper right')
        
        # 2. Rentabilité
        ax2 = plt.subplot(3, 4, 2)
        ebitda_margins = [proj['ebitda_margin'] for proj in calc.projections]
        roic_values = [proj['roic'] for proj in calc.projections]
        
        ax2.plot(years, ebitda_margins, 'r-o', label='Marge EBITDA (%)')
        ax2.plot(years, roic_values, 'purple', marker='s', label='ROIC (%)')
        ax2.axhline(y=calc.assumptions['wacc'], color='orange', linestyle='--', 
                   label=f'WACC ({calc.assumptions["wacc"]:.1f}%)')
        ax2.set_title('Rentabilité', fontweight='bold')
        ax2.legend()
        
        # 3. Pont de valorisation
        ax3 = plt.subplot(3, 4, 3)
        waterfall_data = [
            ('FCF', calc.results['npv_operating_fcf']),
            ('VT', calc.results['terminal_analysis']['terminal_value_pv']),
            ('Liquidités', calc.results['cash_excess']),
            ('Dette', -calc.results['debt_adjustment'])
        ]
        
        x_pos = range(len(waterfall_data))
        values = [item[1] for item in waterfall_data]
        colors = ['skyblue' if v > 0 else 'lightcoral' for v in values]
        
        ax3.bar(x_pos, values, color=colors, alpha=0.7)
        ax3.set_title('Pont de Valorisation', fontweight='bold')
        ax3.set_xticks(x_pos)
        ax3.set_xticklabels([item[0] for item in waterfall_data], rotation=45)
        
        # 4. Scénarios
        ax4 = plt.subplot(3, 4, 4)
        if hasattr(calc, 'scenarios') and calc.scenarios:
            scenario_names = list(calc.scenarios.keys())
            scenario_prices = [calc.scenarios[name]['share_price'] for name in scenario_names]
            
            ax4.bar(scenario_names, scenario_prices, alpha=0.7, 
                   color=['red', 'blue', 'green', 'orange'][:len(scenario_names)])
            ax4.set_title('Prix par Scénario (€)', fontweight='bold')
            ax4.axhline(y=calc.results['share_price'], color='black', linestyle='--')
        
        # 5. Analyse de sensibilité (si disponible)
        ax5 = plt.subplot(3, 4, 5)
        if hasattr(calc, 'sensitivity_analysis') and calc.sensitivity_analysis:
            sens = calc.sensitivity_analysis
            im = ax5.imshow(sens['percent_matrix'], cmap='RdYlGn', aspect='auto')
            ax5.set_title('Sensibilité (%)', fontweight='bold')
            plt.colorbar(im, ax=ax5)
        
        # 6. Distribution Monte Carlo
        ax6 = plt.subplot(3, 4, 6)
        if hasattr(calc, 'monte_carlo_results') and calc.monte_carlo_results:
            mc = calc.monte_carlo_results
            ax6.hist(mc['values'], bins=30, alpha=0.7, color='lightblue')
            ax6.axvline(mc['mean'], color='red', linestyle='--', label=f'Moyenne: €{mc["mean"]:.2f}')
            ax6.set_title('Monte Carlo', fontweight='bold')
            ax6.legend()
        
        # 7. ROIC vs WACC
        ax7 = plt.subplot(3, 4, 7)
        value_creation = [proj['value_creation'] for proj in calc.projections]
        colors_vc = ['green' if vc > 0 else 'red' for vc in value_creation]
        
        ax7.bar(years, value_creation, color=colors_vc, alpha=0.7)
        ax7.axhline(y=0, color='black', linestyle='-')
        ax7.set_title('Création de Valeur', fontweight='bold')
        
        # 8. Investissements
        ax8 = plt.subplot(3, 4, 8)
        capex_values = [proj['total_capex'] for proj in calc.projections]
        ax8.bar(years, capex_values, alpha=0.7, color='brown')
        ax8.set_title('Capex par Année', fontweight='bold')
        
        # 9. Comparaison méthodes
        ax9 = plt.subplot(3, 4, 9)
        if hasattr(calc, 'peer_analysis') and calc.peer_analysis:
            peer = calc.peer_analysis
            methods = ['DCF', 'EV/Rev', 'EV/EBITDA']
            prices = [
                peer['dcf_price'],
                peer['share_prices']['ev_revenue_median'],
                peer['share_prices']['ev_ebitda_median']
            ]
            ax9.bar(methods, prices, alpha=0.7, color=['darkblue', 'lightgreen', 'orange'])
            ax9.set_title('Comparaison Valorisations', fontweight='bold')
        
        # 10. Métriques clés
        ax10 = plt.subplot(3, 4, 10)
        ax10.axis('off')
        summary_text = f"""
RÉSUMÉ EXÉCUTIF

Prix DCF: €{calc.results['share_price']:.2f}
VE: €{calc.results['enterprise_value']:,.0f}M
VE/CA: {calc.results['ev_revenue']:.1f}x
VE/EBITDA: {calc.results['ev_ebitda']:.1f}x

WACC: {calc.assumptions['wacc']:.1f}%
Croissance: {calc.assumptions['terminal_growth']:.1f}%
ROIC: {calc.results['avg_roic']:.1f}%
        """
        
        ax10.text(0.1, 0.9, summary_text, transform=ax10.transAxes, fontsize=10,
                 verticalalignment='top', fontfamily='monospace',
                 bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8))
        
        # 11 & 12. Métriques additionnelles
        for i in range(11, 13):
            ax = plt.subplot(3, 4, i)
            ax.text(0.5, 0.5, f'Graphique {i}\nDisponible en mode\ncomplet', 
                   ha='center', va='center', transform=ax.transAxes)
            ax.set_title(f'Analyse {i}', fontweight='bold')
        
        plt.tight_layout(pad=2.0)
        
        # Sauvegarde
        filename = f"dcf_analysis_{calc.company_name.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        print(f"✅ Graphiques sauvegardés: {filename}")
        
        plt.show()
        return True

class UniversalExporter:
    """Gestionnaire d'export adaptatif"""
    
    def __init__(self):
        self.can_export_excel = HAS_OPENPYXL
    
    def export_results(self, dcf_calculator):
        """Exporte les résultats selon les modules disponibles"""
        if self.can_export_excel:
            return self._export_to_excel(dcf_calculator)
        else:
            return self._export_to_csv(dcf_calculator)
    
    def _export_to_excel(self, calc):
        """Export Excel complet"""
        try:
            wb = Workbook()
            
            # Feuille principale
            ws = wb.active
            ws.title = "Résumé"
            
            ws['A1'] = f"Analyse DCF - {calc.company_name}"
            ws['A1'].font = Font(bold=True, size=16)
            
            # Données principales
            data = [
                ["Prix par Action", f"€{calc.results['share_price']:.2f}"],
                ["Valeur d'Entreprise", f"€{calc.results['enterprise_value']:,.0f}M"],
                ["WACC", f"{calc.assumptions['wacc']:.1f}%"],
                ["Croissance Terminale", f"{calc.assumptions['terminal_growth']:.1f}%"]
            ]
            
            for i, (label, value) in enumerate(data, 3):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
            
            # Feuille projections
            ws_proj = wb.create_sheet("Projections")
            headers = ["Année", "CA", "EBITDA", "FCF", "ROIC"]
            
            for i, header in enumerate(headers, 1):
                ws_proj.cell(1, i, header).font = Font(bold=True)
            
            for row, proj in enumerate(calc.projections, 2):
                ws_proj.cell(row, 1, proj['year'])
                ws_proj.cell(row, 2, proj['revenue'])
                ws_proj.cell(row, 3, proj['ebitda'])
                ws_proj.cell(row, 4, proj['fcf'])
                ws_proj.cell(row, 5, proj['roic'])
            
            filename = f"dcf_analysis_{calc.company_name.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(filename)
            print(f"✅ Export Excel: {filename}")
            return True
            
        except Exception as e:
            print(f"⚠️ Erreur export Excel: {e}")
            return self._export_to_csv(calc)
    
    def _export_to_csv(self, calc):
        """Export CSV de base"""
        try:
            filename = f"dcf_analysis_{calc.company_name.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(f"DCF Analysis - {calc.company_name}\n")
                f.write(f"Date,{datetime.now().strftime('%Y-%m-%d')}\n\n")
                
                f.write("Résumé Exécutif\n")
                f.write(f"Prix par Action,€{calc.results['share_price']:.2f}\n")
                f.write(f"Valeur d'Entreprise,€{calc.results['enterprise_value']:,.0f}M\n")
                f.write(f"WACC,{calc.assumptions['wacc']:.1f}%\n\n")
                
                f.write("Projections\n")
                f.write("Année,CA,EBITDA,FCF,ROIC\n")
                for proj in calc.projections:
                    f.write(f"{proj['year']},{proj['revenue']},{proj['ebitda']},{proj['fcf']},{proj['roic']}\n")
            
            print(f"✅ Export CSV: {filename}")
            return True
            
        except Exception as e:
            print(f"❌ Erreur export CSV: {e}")
            return False

# ==================== CLASSES PRINCIPALES ====================

class MarketDataProvider:
    """Fournisseur de données de marché"""
    
    @staticmethod
    def get_market_data():
        return MARKET_DATA
    
    @staticmethod
    def get_industry_benchmarks(industry):
        return INDUSTRY_BENCHMARKS.get(industry, INDUSTRY_BENCHMARKS["General"])
    
    @staticmethod
    def get_peer_multiples(industry):
        benchmarks = INDUSTRY_BENCHMARKS.get(industry, INDUSTRY_BENCHMARKS["General"])
        ev_revenue_mid = sum(benchmarks["ev_revenue"]) / 2
        ev_ebitda_mid = sum(benchmarks["ev_ebitda"]) / 2
        
        return {
            "ev_revenue_median": ev_revenue_mid,
            "ev_revenue_mean": ev_revenue_mid * 1.05,
            "ev_ebitda_median": ev_ebitda_mid,
            "ev_ebitda_mean": ev_ebitda_mid * 1.08
        }

class WAACCalculator:
    """Calculateur WAAC avec CAPM"""
    
    def __init__(self, risk_free_rate, market_risk_premium):
        self.risk_free_rate = risk_free_rate
        self.market_risk_premium = market_risk_premium
    
    def calculate_cost_of_equity(self, beta, size_premium=0):
        return self.risk_free_rate + beta * self.market_risk_premium + size_premium
    
    def calculate_cost_of_debt(self, credit_spread):
        return self.risk_free_rate + credit_spread
    
    def calculate_wacc(self, market_value_equity, market_value_debt,
                      cost_of_equity, cost_of_debt, tax_rate):
        total_value = market_value_equity + market_value_debt
        
        if total_value == 0:
            return cost_of_equity
        
        weight_equity = market_value_equity / total_value
        weight_debt = market_value_debt / total_value
        after_tax_cost_debt = cost_of_debt * (1 - tax_rate / 100)
        
        return weight_equity * cost_of_equity + weight_debt * after_tax_cost_debt

class GrowthModel:
    """Modèle de croissance multi-phases"""
    
    @staticmethod
    def three_phase_growth(high_growth_years, transition_years,
                          high_growth_rate, terminal_growth_rate, total_years):
        growth_rates = []
        
        # Phase 1: Forte croissance
        for year in range(min(high_growth_years, total_years)):
            growth_rates.append(high_growth_rate)
        
        # Phase 2: Transition
        if total_years > high_growth_years:
            transition_start = len(growth_rates)
            transition_end = min(high_growth_years + transition_years, total_years)
            
            for year in range(transition_start, transition_end):
                if transition_years > 1:
                    progress = (year - transition_start) / (transition_years - 1)
                else:
                    progress = 1
                rate = high_growth_rate - progress * (high_growth_rate - terminal_growth_rate)
                growth_rates.append(max(rate, terminal_growth_rate))
        
        # Phase 3: Croissance terminale
        while len(growth_rates) < total_years:
            growth_rates.append(terminal_growth_rate)
        
        return growth_rates[:total_years]

class MonteCarloSimulator:
    """Simulateur Monte Carlo adaptatif"""
    
    def __init__(self, iterations=5000):
        self.iterations = iterations
    
    def simulate_dcf(self, base_assumptions, sensitivity_params):
        results = []
        
        for _ in range(self.iterations):
            assumptions = base_assumptions.copy()
            
            for param, (distribution, *args) in sensitivity_params.items():
                if distribution == "normal":
                    mean, std = args
                    assumptions[param] = random.gauss(mean, std)
                elif distribution == "uniform":
                    low, high = args
                    assumptions[param] = random.uniform(low, high)
                elif distribution == "triangular":
                    low, high, mode = args
                    assumptions[param] = random.triangular(low, high, mode)
            
            simulated_value = self._quick_dcf_calc(assumptions)
            results.append(simulated_value)
        
        return {
            "mean": statistics.mean(results),
            "median": statistics.median(results),
            "std": statistics.stdev(results) if len(results) > 1 else 0,
            "percentile_5": np.percentile(results, 5),
            "percentile_25": np.percentile(results, 25),
            "percentile_75": np.percentile(results, 75),
            "percentile_95": np.percentile(results, 95),
            "values": results
        }
    
    def _quick_dcf_calc(self, assumptions):
        base_fcf = assumptions.get("base_fcf", 100)
        growth = assumptions.get("terminal_growth", 2.5) / 100
        wacc = assumptions.get("wacc", 10) / 100
        
        if wacc <= growth:
            return 0
        
        terminal_value = base_fcf * (1 + growth) / (wacc - growth)
        return terminal_value / (1 + wacc) ** 5

class UniversalDCFCalculator:
    """Calculateur DCF universel et adaptatif"""
    
    def __init__(self):
        self.company_name = ""
        self.industry = "General"
        self.assumptions = {}
        self.projections = []
        self.results = {}
        self.scenarios = {}
        self.sensitivity_analysis = {}
        self.monte_carlo_results = {}
        self.peer_analysis = {}
        
        # Composants adaptatifs
        market_data = MarketDataProvider.get_market_data()
        self.wacc_calculator = WAACCalculator(
            market_data["risk_free_rate"],
            market_data["market_risk_premium"]
        )
        self.monte_carlo = MonteCarloSimulator()
        self.visualizer = UniversalVisualizer()
        self.exporter = UniversalExporter()
        
    def display_banner(self):
        """Bannière avec détection de capacités"""
        print("=" * 80)
        print("🚀 DCF CALCULATOR UNIVERSAL - COMPATIBLE TOUS ENVIRONNEMENTS")
        print("📊 S'adapte automatiquement aux modules Python disponibles")
        print("🎯 Fonctionne avec Python seul ou avec modules avancés")
        print(f"📅 Version 2.0 Universal - {datetime.now().strftime('%Y-%m-%d')}")
        print("=" * 80)
        
        # Affichage des capacités
        print(f"\n🔧 CAPACITÉS DÉTECTÉES:")
        print(f"   Graphiques avancés:    {'✅ OUI' if HAS_MATPLOTLIB else '❌ NON (pip install matplotlib)'}")
        print(f"   Export Excel:          {'✅ OUI' if HAS_OPENPYXL else '❌ NON (pip install openpyxl)'}")
        print(f"   Tableaux formatés:     {'✅ OUI' if HAS_TABULATE else '❌ NON (pip install tabulate)'}")
        print(f"   Calculs optimisés:     {'✅ OUI' if HAS_NUMPY else '❌ NON (pip install numpy)'}")
        print(f"   Mode de fonctionnement: {'🚀 COMPLET' if all([HAS_MATPLOTLIB, HAS_OPENPYXL, HAS_TABULATE]) else '⚡ ADAPTATIF'}")
        print()
    
    def get_detailed_assumptions(self):
        """Collecte d'hypothèses avec interface adaptative"""
        print("💰 HYPOTHÈSES FINANCIÈRES DÉTAILLÉES")
        print("-" * 50)
        
        benchmarks = MarketDataProvider.get_industry_benchmarks(self.industry)
        market_data = MarketDataProvider.get_market_data()
        
        # Affichage des benchmarks
        print(f"📊 Benchmarks {self.industry.upper()}:")
        print(f"  EBITDA Margin:     {benchmarks['ebitda_margin'][0]:.1f}% - {benchmarks['ebitda_margin'][1]:.1f}%")
        print(f"  Revenue Growth:    {benchmarks['revenue_growth'][0]:.1f}% - {benchmarks['revenue_growth'][1]:.1f}%")
        print(f"  Beta:              {benchmarks['beta'][0]:.1f} - {benchmarks['beta'][1]:.1f}")
        print(f"  ROIC:              {benchmarks['roic'][0]:.1f}% - {benchmarks['roic'][1]:.1f}%")
        print()
        
        try:
            # Données de base
            current_revenue = float(input("📈 Chiffre d'affaires actuel (M€): ") or "500")
            current_ebitda = float(input("💎 EBITDA actuel (M€): ") or str(current_revenue * 0.15))
            
            # Modèle de croissance
            print("\n🚀 MODÈLE DE CROISSANCE")
            print("1. Croissance uniforme (simple)")
            print("2. Modèle 3 phases (professionnel)")
            
            growth_model_choice = input("Choix du modèle [2]: ") or "2"
            
            if growth_model_choice == "2":
                high_growth_years = int(input("Années forte croissance [3]: ") or "3")
                transition_years = int(input("Années de transition [2]: ") or "2")
                high_growth_rate = float(input("Taux forte croissance (%) [15]: ") or "15")
                terminal_growth = float(input("Croissance perpétuelle (%) [2.5]: ") or "2.5")
                
                total_years = 10
                revenue_growth_rates = GrowthModel.three_phase_growth(
                    high_growth_years, transition_years, high_growth_rate, terminal_growth, total_years
                )
            else:
                total_years = 5
                avg_growth = float(input("Croissance moyenne (%) [8]: ") or "8")
                revenue_growth_rates = [avg_growth] * total_years
                terminal_growth = float(input("Croissance perpétuelle (%) [2.5]: ") or "2.5")
            
            # Paramètres de profitabilité
            print("\n💰 PARAMÈTRES DE PROFITABILITÉ")
            ebitda_margin_start = float(input(f"Marge EBITDA initiale (%) [{(current_ebitda/current_revenue)*100:.1f}]: ") 
                                       or str((current_ebitda/current_revenue)*100))
            ebitda_margin_terminal = float(input(f"Marge EBITDA terminale (%) [{ebitda_margin_start:.1f}]: ") 
                                          or str(ebitda_margin_start))
            
            # Structure de capital et WACC
            print("\n🏦 STRUCTURE DE CAPITAL ET WAAC")
            print(f"Taux sans risque: {market_data['risk_free_rate']:.1f}%")
            print(f"Prime de risque marché: {market_data['market_risk_premium']:.1f}%")
            
            beta = float(input(f"Beta [{sum(benchmarks['beta'])/2:.1f}]: ") or str(sum(benchmarks['beta'])/2))
            size_premium = float(input("Prime de taille (%) [0.5]: ") or "0.5")
            credit_spread = float(input("Spread crédit (%) [2.0]: ") or "2.0")
            
            # Calculs WACC
            cost_of_equity = self.wacc_calculator.calculate_cost_of_equity(beta, size_premium)
            cost_of_debt = self.wacc_calculator.calculate_cost_of_debt(credit_spread)
            
            print(f"→ Coût des fonds propres: {cost_of_equity:.2f}%")
            print(f"→ Coût de la dette: {cost_of_debt:.2f}%")
            
            # Structure financière
            market_value_equity = float(input("Valeur marché fonds propres (M€) [800]: ") or "800")
            market_value_debt = float(input("Valeur marché dette (M€) [200]: ") or "200")
            tax_rate = float(input("Taux d'imposition (%) [25]: ") or "25")
            
            wacc = self.wacc_calculator.calculate_wacc(
                market_value_equity, market_value_debt, cost_of_equity, cost_of_debt, tax_rate
            )
            
            print(f"→ WAAC calculé: {wacc:.2f}%")
            
            # Investissements
            print("\n🏗️ MODÈLE D'INVESTISSEMENTS")
            capex_base = float(input("Capex maintenance (% CA) [3]: ") or "3")
            capex_growth = float(input("Capex croissance (% Δ CA) [25]: ") or "25")
            nwc_percent = float(input("BFR (% CA) [12]: ") or "12")
            depreciation_rate = float(input("Taux dépréciation (% CA) [4]: ") or "4")
            
            # Validation
            if wacc <= terminal_growth:
                print(f"❌ ERREUR: WAAC ({wacc:.1f}%) doit être > Croissance perpétuelle ({terminal_growth:.1f}%)")
                return False
            
            # Stockage
            self.assumptions = {
                'current_revenue': current_revenue,
                'current_ebitda': current_ebitda,
                'revenue_growth_rates': revenue_growth_rates,
                'total_years': total_years,
                'ebitda_margin_start': ebitda_margin_start,
                'ebitda_margin_terminal': ebitda_margin_terminal,
                'terminal_growth': terminal_growth,
                'beta': beta,
                'size_premium': size_premium,
                'cost_of_equity': cost_of_equity,
                'cost_of_debt': cost_of_debt,
                'wacc': wacc,
                'market_value_equity': market_value_equity,
                'market_value_debt': market_value_debt,
                'tax_rate': tax_rate,
                'capex_base': capex_base,
                'capex_growth': capex_growth,
                'nwc_percent': nwc_percent,
                'depreciation_rate': depreciation_rate,
                'growth_model': growth_model_choice
            }
            
            print("\n✅ Hypothèses validées!")
            return True
            
        except ValueError:
            print("❌ Erreur: Veuillez entrer des valeurs numériques valides")
            return False
    
    def calculate_advanced_projections(self):
        """Calcule les projections multi-phases"""
        print("\n🔄 CALCUL DES PROJECTIONS AVANCÉES...")
        
        projections = []
        invested_capital = self.assumptions['current_revenue'] * 0.8
        previous_revenue = self.assumptions['current_revenue']
        
        for year in range(1, self.assumptions['total_years'] + 1):
            # Croissance
            growth_rate = self.assumptions['revenue_growth_rates'][year - 1] / 100
            revenue = previous_revenue * (1 + growth_rate)
            revenue_increase = revenue - previous_revenue
            
            # Évolution marge EBITDA
            margin_progress = (year - 1) / max(1, self.assumptions['total_years'] - 1)
            ebitda_margin = (
                self.assumptions['ebitda_margin_start'] + 
                margin_progress * (self.assumptions['ebitda_margin_terminal'] - self.assumptions['ebitda_margin_start'])
            )
            
            # Cascade financière
            ebitda = revenue * (ebitda_margin / 100)
            depreciation = revenue * (self.assumptions['depreciation_rate'] / 100)
            ebit = ebitda - depreciation
            tax = ebit * (self.assumptions['tax_rate'] / 100)
            nopat = ebit - tax
            
            # Investissements
            capex_maintenance = revenue * (self.assumptions['capex_base'] / 100)
            capex_growth = revenue_increase * (self.assumptions['capex_growth'] / 100)
            total_capex = capex_maintenance + capex_growth
            
            # BFR
            if year == 1:
                delta_nwc = revenue * (self.assumptions['nwc_percent'] / 100)
            else:
                delta_nwc = revenue_increase * (self.assumptions['nwc_percent'] / 100)
            
            # FCF
            fcf = nopat + depreciation - total_capex - delta_nwc
            
            # ROIC
            invested_capital += total_capex + delta_nwc - depreciation
            roic = (nopat / invested_capital * 100) if invested_capital > 0 else 0
            value_creation = roic - self.assumptions['wacc']
            
            # Actualisation
            discount_factor = (1 + self.assumptions['wacc'] / 100) ** year
            present_value_fcf = fcf / discount_factor
            
            projection = {
                'year': year,
                'revenue': round(revenue, 1),
                'revenue_growth': round(growth_rate * 100, 1),
                'revenue_increase': round(revenue_increase, 1),
                'ebitda': round(ebitda, 1),
                'ebitda_margin': round(ebitda_margin, 1),
                'depreciation': round(depreciation, 1),
                'ebit': round(ebit, 1),
                'tax': round(tax, 1),
                'nopat': round(nopat, 1),
                'capex_maintenance': round(capex_maintenance, 1),
                'capex_growth': round(capex_growth, 1),
                'total_capex': round(total_capex, 1),
                'delta_nwc': round(delta_nwc, 1),
                'fcf': round(fcf, 1),
                'invested_capital': round(invested_capital, 1),
                'roic': round(roic, 1),
                'value_creation': round(value_creation, 1),
                'discount_factor': round(discount_factor, 4),
                'pv_fcf': round(present_value_fcf, 1)
            }
            
            projections.append(projection)
            previous_revenue = revenue
        
        self.projections = projections
        print("✅ Projections calculées")
    
    def calculate_terminal_value_advanced(self):
        """Calcul valeur terminale avec validation"""
        final_projection = self.projections[-1]
        terminal_fcf = final_projection['fcf'] * (1 + self.assumptions['terminal_growth'] / 100)
        
        wacc_rate = self.assumptions['wacc'] / 100
        terminal_growth_rate = self.assumptions['terminal_growth'] / 100
        
        # Gordon
        gordon_terminal_value = terminal_fcf / (wacc_rate - terminal_growth_rate)
        
        # Multiple de sortie
        terminal_ebitda = final_projection['ebitda'] * (1 + self.assumptions['terminal_growth'] / 100)
        benchmarks = MarketDataProvider.get_industry_benchmarks(self.industry)
        exit_multiple = sum(benchmarks['ev_ebitda']) / 2
        multiple_terminal_value = terminal_ebitda * exit_multiple
        
        # Valeur conservative
        terminal_value = min(gordon_terminal_value, multiple_terminal_value * 1.1)
        
        # Actualisation
        discount_factor = (1 + wacc_rate) ** self.assumptions['total_years']
        terminal_value_pv = terminal_value / discount_factor
        
        return {
            'terminal_fcf': round(terminal_fcf, 1),
            'gordon_terminal_value': round(gordon_terminal_value, 0),
            'multiple_terminal_value': round(multiple_terminal_value, 0),
            'terminal_value_used': round(terminal_value, 0),
            'terminal_value_pv': round(terminal_value_pv, 0),
            'exit_multiple': round(exit_multiple, 1),
            'method_used': 'Conservative (Min of Gordon vs Multiple)'
        }
    
    def calculate_dcf_valuation_advanced(self):
        """Calcul DCF complet"""
        print("💎 CALCUL DCF AVANCÉ...")
        
        # FCF actualisés
        npv_operating_fcf = sum([proj['pv_fcf'] for proj in self.projections])
        
        # Valeur terminale
        terminal_analysis = self.calculate_terminal_value_advanced()
        
        # Valeur d'entreprise
        enterprise_value = npv_operating_fcf + terminal_analysis['terminal_value_pv']
        
        # Ajustements
        cash_excess = float(input("\n💰 Liquidités excédentaires (M€) [0]: ") or "0")
        debt_book_value = self.assumptions['market_value_debt']
        minority_interests = float(input("👥 Intérêts minoritaires (M€) [0]: ") or "0")
        
        # Valeur fonds propres
        equity_value = enterprise_value + cash_excess - debt_book_value - minority_interests
        
        # Prix par action
        shares_outstanding = float(input("📊 Nombre d'actions (M) [25]: ") or "25")
        share_price = equity_value / shares_outstanding
        
        # Multiples
        current_revenue = self.assumptions['current_revenue']
        current_ebitda = self.assumptions['current_ebitda']
        
        ev_revenue = enterprise_value / current_revenue
        ev_ebitda = enterprise_value / current_ebitda
        
        # Pourcentage valeur terminale
        terminal_percentage = (terminal_analysis['terminal_value_pv'] / enterprise_value) * 100
        
        # Métriques de qualité
        avg_roic = sum([proj['roic'] for proj in self.projections]) / len(self.projections)
        avg_value_creation = sum([proj['value_creation'] for proj in self.projections]) / len(self.projections)
        
        self.results = {
            'npv_operating_fcf': round(npv_operating_fcf, 0),
            'terminal_analysis': terminal_analysis,
            'enterprise_value': round(enterprise_value, 0),
            'cash_excess': round(cash_excess, 0),
            'debt_adjustment': round(debt_book_value, 0),
            'minority_interests': round(minority_interests, 0),
            'equity_value': round(equity_value, 0),
            'shares_outstanding': round(shares_outstanding, 1),
            'share_price': round(share_price, 2),
            'ev_revenue': round(ev_revenue, 1),
            'ev_ebitda': round(ev_ebitda, 1),
            'terminal_percentage': round(terminal_percentage, 1),
            'avg_roic': round(avg_roic, 1),
            'avg_value_creation': round(avg_value_creation, 1),
            'wacc_spread': round(avg_roic - self.assumptions['wacc'], 1)
        }
        
        print("✅ DCF calculé")
    
    def sensitivity_analysis_2d(self):
        """Analyse de sensibilité 2D"""
        print("📊 ANALYSE DE SENSIBILITÉ 2D...")
        
        base_wacc = self.assumptions['wacc']
        base_terminal_growth = self.assumptions['terminal_growth']
        base_share_price = self.results['share_price']
        
        # Ranges
        wacc_range = np.arange(base_wacc - 2, base_wacc + 2.5, 0.5)
        growth_range = np.arange(max(0.5, base_terminal_growth - 1.5), base_terminal_growth + 1.5, 0.5)
        
        sensitivity_matrix = []
        sensitivity_data = {}
        
        for wacc_test in wacc_range:
            row = []
            for growth_test in growth_range:
                if wacc_test <= growth_test:
                    price = 0
                else:
                    # Recalcul rapide
                    final_fcf = self.projections[-1]['fcf']
                    terminal_fcf = final_fcf * (1 + growth_test / 100)
                    terminal_value = terminal_fcf / (wacc_test / 100 - growth_test / 100)
                    terminal_value_pv = terminal_value / ((1 + wacc_test / 100) ** self.assumptions['total_years'])
                    
                    npv_fcf = sum([proj['fcf'] / ((1 + wacc_test / 100) ** proj['year']) for proj in self.projections])
                    
                    enterprise_value = npv_fcf + terminal_value_pv
                    equity_value = enterprise_value - self.results['debt_adjustment'] + self.results['cash_excess'] - self.results['minority_interests']
                    price = equity_value / self.results['shares_outstanding']
                
                row.append(round(price, 2))
            sensitivity_matrix.append(row)
        
        # Variations en %
        sensitivity_percent = []
        for i, wacc_test in enumerate(wacc_range):
            row_percent = []
            for j, growth_test in enumerate(growth_range):
                if sensitivity_matrix[i][j] > 0:
                    variation = (sensitivity_matrix[i][j] / base_share_price - 1) * 100
                    row_percent.append(round(variation, 1))
                else:
                    row_percent.append("N/A")
            sensitivity_percent.append(row_percent)
        
        self.sensitivity_analysis = {
            'wacc_range': wacc_range,
            'growth_range': growth_range,
            'price_matrix': sensitivity_matrix,
            'percent_matrix': sensitivity_percent,
            'base_wacc': base_wacc,
            'base_growth': base_terminal_growth,
            'base_price': base_share_price
        }
        
        print("✅ Sensibilité 2D calculée")
    
    def monte_carlo_simulation(self):
        """Simulation Monte Carlo"""
        print("🎲 SIMULATION MONTE CARLO...")
        
        sensitivity_params = {
            'wacc': ('normal', self.assumptions['wacc'], 1.0),
            'terminal_growth': ('triangular', 
                              max(0.5, self.assumptions['terminal_growth'] - 1), 
                              min(4.0, self.assumptions['terminal_growth'] + 1), 
                              self.assumptions['terminal_growth']),
            'ebitda_margin': ('normal', self.assumptions['ebitda_margin_terminal'], 2.0),
            'base_fcf': ('normal', self.projections[-1]['fcf'], self.projections[-1]['fcf'] * 0.2)
        }
        
        self.monte_carlo_results = self.monte_carlo.simulate_dcf(self.assumptions, sensitivity_params)
        
        # Analyse des résultats
        base_price = self.results['share_price']
        mc_results = self.monte_carlo_results
        
        prob_positive = len([x for x in mc_results['values'] if x > base_price]) / len(mc_results['values']) * 100
        prob_above_150 = len([x for x in mc_results['values'] if x > base_price * 1.5]) / len(mc_results['values']) * 100
        prob_below_50 = len([x for x in mc_results['values'] if x < base_price * 0.5]) / len(mc_results['values']) * 100
        
        mc_results.update({
            'base_price': base_price,
            'prob_positive': round(prob_positive, 1),
            'prob_above_150': round(prob_above_150, 1),
            'prob_below_50': round(prob_below_50, 1),
            'var_95': round(mc_results['percentile_5'], 2),
            'var_99': round(np.percentile(mc_results['values'], 1), 2)
        })
        
        print("✅ Monte Carlo terminé")
    
    def comparative_valuation(self):
        """Valorisation comparative"""
        print("📊 VALORISATION COMPARATIVE...")
        
        peer_multiples = MarketDataProvider.get_peer_multiples(self.industry)
        current_revenue = self.assumptions['current_revenue']
        current_ebitda = self.assumptions['current_ebitda']
        
        # Valorisations par multiples
        valuations = {
            'ev_revenue_median': current_revenue * peer_multiples['ev_revenue_median'],
            'ev_revenue_mean': current_revenue * peer_multiples['ev_revenue_mean'],
            'ev_ebitda_median': current_ebitda * peer_multiples['ev_ebitda_median'],
            'ev_ebitda_mean': current_ebitda * peer_multiples['ev_ebitda_mean']
        }
        
        # Conversion en prix par action
        def ev_to_share_price(ev):
            equity_value = ev - self.results['debt_adjustment'] + self.results['cash_excess']
            return equity_value / self.results['shares_outstanding']
        
        share_prices = {
            key: round(ev_to_share_price(ev), 2) for key, ev in valuations.items()
        }
        
        # Comparaisons
        dcf_price = self.results['share_price']
        comparisons = {}
        for method, price in share_prices.items():
            premium_discount = (dcf_price / price - 1) * 100 if price > 0 else 0
            comparisons[method] = {
                'price': price,
                'premium_discount': round(premium_discount, 1)
            }
        
        # Moyenne pondérée
        avg_ev_revenue = (peer_multiples['ev_revenue_median'] + peer_multiples['ev_revenue_mean']) / 2
        avg_ev_ebitda = (peer_multiples['ev_ebitda_median'] + peer_multiples['ev_ebitda_mean']) / 2
        
        blended_ev = (current_revenue * avg_ev_revenue + current_ebitda * avg_ev_ebitda) / 2
        blended_price = ev_to_share_price(blended_ev)
        
        self.peer_analysis = {
            'peer_multiples': peer_multiples,
            'valuations_ev': valuations,
            'share_prices': share_prices,
            'comparisons': comparisons,
            'blended_price': round(blended_price, 2),
            'dcf_vs_blended': round((dcf_price / blended_price - 1) * 100, 1),
            'dcf_price': dcf_price
        }
        
        print("✅ Valorisation comparative terminée")
    
    def advanced_scenario_analysis(self):
        """Analyse de scénarios sophistiquée"""
        print("🎯 ANALYSE DE SCÉNARIOS AVANCÉE...")
        
        scenarios_config = {
            'Bear Case': {
                'revenue_adj': -0.30, 'margin_adj': -0.20, 'wacc_adj': 2.0,
                'terminal_growth_adj': -0.5, 'description': 'Récession + Concurrence'
            },
            'Base Case': {
                'revenue_adj': 0.0, 'margin_adj': 0.0, 'wacc_adj': 0.0,
                'terminal_growth_adj': 0.0, 'description': 'Consensus Marché'
            },
            'Bull Case': {
                'revenue_adj': 0.25, 'margin_adj': 0.15, 'wacc_adj': -1.0,
                'terminal_growth_adj': 0.5, 'description': 'Expansion + Synergies'
            },
            'Stress Test': {
                'revenue_adj': -0.50, 'margin_adj': -0.30, 'wacc_adj': 3.0,
                'terminal_growth_adj': -1.0, 'description': 'Crise Majeure'
            }
        }
        
        scenario_results = {}
        original_assumptions = self.assumptions.copy()
        
        for scenario_name, adjustments in scenarios_config.items():
            test_assumptions = original_assumptions.copy()
            
            # Ajustements
            test_assumptions['revenue_growth_rates'] = [
                max(-20, g * (1 + adjustments['revenue_adj'])) 
                for g in original_assumptions['revenue_growth_rates']
            ]
            
            test_assumptions['ebitda_margin_terminal'] = max(
                5, original_assumptions['ebitda_margin_terminal'] * (1 + adjustments['margin_adj'])
            )
            
            test_assumptions['wacc'] = original_assumptions['wacc'] + adjustments['wacc_adj']
            test_assumptions['terminal_growth'] = max(
                0, original_assumptions['terminal_growth'] + adjustments['terminal_growth_adj']
            )
            
            # Validation
            if test_assumptions['wacc'] <= test_assumptions['terminal_growth']:
                test_assumptions['wacc'] = test_assumptions['terminal_growth'] + 1.0
            
            # Calcul du scénario
            scenario_price = self._calculate_scenario_price(test_assumptions)
            
            scenario_results[scenario_name] = {
                'share_price': round(scenario_price, 2),
                'vs_base': round((scenario_price / self.results['share_price'] - 1) * 100, 1),
                'description': adjustments['description'],
                'assumptions': {
                    'avg_growth': round(sum(test_assumptions['revenue_growth_rates']) / len(test_assumptions['revenue_growth_rates']), 1),
                    'ebitda_margin': round(test_assumptions['ebitda_margin_terminal'], 1),
                    'wacc': round(test_assumptions['wacc'], 1),
                    'terminal_growth': round(test_assumptions['terminal_growth'], 1)
                }
            }
        
        self.scenarios = scenario_results
        print("✅ Scénarios calculés")
    
    def _calculate_scenario_price(self, test_assumptions):
        """Calcul rapide du prix pour un scénario"""
        revenue = test_assumptions['current_revenue']
        total_fcf = 0
        
        for year in range(1, test_assumptions['total_years'] + 1):
            growth_rate = test_assumptions['revenue_growth_rates'][year - 1] / 100
            revenue *= (1 + growth_rate)
            
            ebitda = revenue * (test_assumptions['ebitda_margin_terminal'] / 100)
            nopat = ebitda * 0.7  # Approximation
            fcf = nopat * 0.8  # Approximation
            
            pv_fcf = fcf / ((1 + test_assumptions['wacc'] / 100) ** year)
            total_fcf += pv_fcf
        
        # Valeur terminale
        final_fcf = fcf * (1 + test_assumptions['terminal_growth'] / 100)
        terminal_value = final_fcf / (test_assumptions['wacc'] / 100 - test_assumptions['terminal_growth'] / 100)
        terminal_pv = terminal_value / ((1 + test_assumptions['wacc'] / 100) ** test_assumptions['total_years'])
        
        # Prix
        enterprise_value = total_fcf + terminal_pv
        equity_value = enterprise_value - self.results['debt_adjustment'] + self.results['cash_excess']
        return equity_value / self.results['shares_outstanding']
    
    def display_professional_results(self):
        """Affichage des résultats adaptatif"""
        print("\n" + "=" * 100)
        print(f"📊 ANALYSE DCF UNIVERSELLE - {self.company_name.upper()}")
        print(f"🏭 Secteur: {self.industry} | 📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 100)
        
        # Résumé exécutif
        print(f"\n💎 RÉSUMÉ EXÉCUTIF")
        print("-" * 50)
        
        executive_data = [
            ["Prix par Action", f"€{self.results['share_price']:.2f}"],
            ["Valeur d'Entreprise", f"€{self.results['enterprise_value']:,.0f}M"],
            ["Valeur Fonds Propres", f"€{self.results['equity_value']:,.0f}M"],
            ["Multiple VE/CA", f"{self.results['ev_revenue']:.1f}x"],
            ["Multiple VE/EBITDA", f"{self.results['ev_ebitda']:.1f}x"],
            ["% Valeur Terminale", f"{self.results['terminal_percentage']:.1f}%"],
            ["ROIC Moyen", f"{self.results['avg_roic']:.1f}%"],
            ["Création de Valeur", f"{self.results['wacc_spread']:+.1f}%"]
        ]
        
        print(SimpleTable.create_table(executive_data, ["Métrique", "Valeur"]))
        
        # Validation vs benchmarks
        benchmarks = MarketDataProvider.get_industry_benchmarks(self.industry)
        print(f"\n📊 VALIDATION VS BENCHMARKS {self.industry.upper()}")
        print("-" * 60)
        
        def get_status(value, range_tuple):
            return "✅" if range_tuple[0] <= value <= range_tuple[1] else "⚠️"
        
        validation_data = [
            ["VE/Chiffre d'Affaires", f"{self.results['ev_revenue']:.1f}x", 
             f"[{benchmarks['ev_revenue'][0]:.1f}x - {benchmarks['ev_revenue'][1]:.1f}x]",
             get_status(self.results['ev_revenue'], benchmarks['ev_revenue'])],
            ["VE/EBITDA", f"{self.results['ev_ebitda']:.1f}x",
             f"[{benchmarks['ev_ebitda'][0]:.1f}x - {benchmarks['ev_ebitda'][1]:.1f}x]",
             get_status(self.results['ev_ebitda'], benchmarks['ev_ebitda'])],
            ["ROIC", f"{self.results['avg_roic']:.1f}%",
             f"[{benchmarks['roic'][0]:.1f}% - {benchmarks['roic'][1]:.1f}%]",
             get_status(self.results['avg_roic'], benchmarks['roic'])],
            ["WACC", f"{self.assumptions['wacc']:.1f}%",
             f"[{benchmarks['wacc'][0]:.1f}% - {benchmarks['wacc'][1]:.1f}%]",
             get_status(self.assumptions['wacc'], benchmarks['wacc'])]
        ]
        
        print(SimpleTable.create_table(validation_data, ["Métrique", "Valeur", "Benchmark", "Status"]))
        
        # Projections financières
        print(f"\n📈 PROJECTIONS FINANCIÈRES DÉTAILLÉES (€M)")
        print("-" * 80)
        
        proj_data = []
        for proj in self.projections:
            row = [
                proj['year'],
                f"{proj['revenue']:.0f}",
                f"{proj['revenue_growth']:+.1f}%",
                f"{proj['ebitda']:.0f}",
                f"{proj['ebitda_margin']:.1f}%",
                f"{proj['fcf']:.0f}",
                f"{proj['roic']:.1f}%"
            ]
            proj_data.append(row)
        
        headers = ["An", "CA", "Δ%", "EBITDA", "Marge%", "FCF", "ROIC%"]
        print(SimpleTable.create_table(proj_data, headers))
        
        # Pont de valorisation
        print(f"\n💰 PONT DE VALORISATION DÉTAILLÉ")
        print("-" * 50)
        
        pont_data = [
            ["Valeur Actuelle FCF", f"€{self.results['npv_operating_fcf']:,.0f}M"],
            ["VA Valeur Terminale", f"€{self.results['terminal_analysis']['terminal_value_pv']:,.0f}M"],
            ["VALEUR D'ENTREPRISE", f"€{self.results['enterprise_value']:,.0f}M"],
            ["Plus: Liquidités", f"€{self.results['cash_excess']:,.0f}M"],
            ["Moins: Dette Nette", f"€{self.results['debt_adjustment']:,.0f}M"],
            ["VALEUR FONDS PROPRES", f"€{self.results['equity_value']:,.0f}M"],
            ["Nombre d'Actions", f"{self.results['shares_outstanding']:.1f}M"],
            ["💎 PRIX PAR ACTION", f"€{self.results['share_price']:.2f}"]
        ]
        
        print(SimpleTable.create_table(pont_data, ["Composante", "Montant"]))
        
        # Analyse de scénarios
        if hasattr(self, 'scenarios') and self.scenarios:
            print(f"\n🎯 ANALYSE DE SCÉNARIOS")
            print("-" * 50)
            
            scenario_data = []
            for name, scenario in self.scenarios.items():
                row = [
                    name,
                    scenario['description'],
                    f"€{scenario['share_price']:.2f}",
                    f"{scenario['vs_base']:+.1f}%"
                ]
                scenario_data.append(row)
            
            headers_scenario = ["Scénario", "Description", "Prix", "vs Base"]
            print(SimpleTable.create_table(scenario_data, headers_scenario))
        
        # Analyse de sensibilité 2D
        if hasattr(self, 'sensitivity_analysis') and self.sensitivity_analysis:
            print(f"\n📊 ANALYSE DE SENSIBILITÉ 2D - PRIX PAR ACTION (€)")
            print("-" * 70)
            
            sens = self.sensitivity_analysis
            
            headers_sens = ['WACC\\Growth'] + [f'{g:.1f}%' for g in sens['growth_range']]
            sens_table_data = []
            
            for i, wacc_val in enumerate(sens['wacc_range']):
                row = [f'{wacc_val:.1f}%'] + [f'€{price:.2f}' if price > 0 else 'N/A' 
                                              for price in sens['price_matrix'][i]]
                sens_table_data.append(row)
            
            print(SimpleTable.create_table(sens_table_data, headers_sens))
        
        # Monte Carlo
        if hasattr(self, 'monte_carlo_results') and self.monte_carlo_results:
            print(f"\n🎲 SIMULATION MONTE CARLO - ANALYSE DE RISQUE")
            print("-" * 60)
            
            mc = self.monte_carlo_results
            mc_data = [
                ["Moyenne", f"€{mc['mean']:.2f}"],
                ["Médiane", f"€{mc['median']:.2f}"],
                ["Écart-Type", f"€{mc['std']:.2f}"],
                ["Percentile 5% (VaR 95%)", f"€{mc['percentile_5']:.2f}"],
                ["Percentile 95%", f"€{mc['percentile_95']:.2f}"],
                ["Prob. Prix > Base", f"{mc['prob_positive']:.1f}%"],
                ["Prob. Prix > +50%", f"{mc['prob_above_150']:.1f}%"]
            ]
            
            print(SimpleTable.create_table(mc_data, ["Métrique", "Valeur"]))
        
        # Valorisation comparative
        if hasattr(self, 'peer_analysis') and self.peer_analysis:
            print(f"\n📊 VALORISATION COMPARATIVE")
            print("-" * 50)
            
            peer = self.peer_analysis
            comp_data = [
                ["DCF (Notre Modèle)", f"€{peer['dcf_price']:.2f}", "Base"],
                ["EV/Revenue Médian", f"€{peer['share_prices']['ev_revenue_median']:.2f}", 
                 f"{peer['comparisons']['ev_revenue_median']['premium_discount']:+.1f}%"],
                ["EV/EBITDA Médian", f"€{peer['share_prices']['ev_ebitda_median']:.2f}",
                 f"{peer['comparisons']['ev_ebitda_median']['premium_discount']:+.1f}%"],
                ["Moyenne Pondérée", f"€{peer['blended_price']:.2f}",
                 f"{peer['dcf_vs_blended']:+.1f}%"]
            ]
            
            headers_comp = ["Méthode", "Prix", "DCF Premium/(Discount)"]
            print(SimpleTable.create_table(comp_data, headers_comp))
        
        # Analyse de qualité
        print(f"\n⚠️ ANALYSE DE QUALITÉ ET RECOMMANDATIONS")
        print("-" * 60)
        
        quality_indicators = []
        
        # Tests de qualité
        if self.results['avg_roic'] > self.assumptions['wacc']:
            quality_indicators.append("✅ ROIC > WACC : Création de valeur positive")
        else:
            quality_indicators.append("❌ ROIC < WACC : Destruction de valeur")
        
        if self.results['terminal_percentage'] < 70:
            quality_indicators.append("✅ Valeur terminale < 70% : Modèle robuste")
        else:
            quality_indicators.append("⚠️ Valeur terminale > 70% : Sensibilité élevée")
        
        if benchmarks['ev_ebitda'][0] <= self.results['ev_ebitda'] <= benchmarks['ev_ebitda'][1]:
            quality_indicators.append("✅ Multiples cohérents avec l'industrie")
        else:
            quality_indicators.append("⚠️ Multiples hors benchmark industrie")
        
        if self.assumptions['terminal_growth'] < self.assumptions['wacc'] - 3:
            quality_indicators.append("✅ Écart WACC-Croissance > 3% : Modèle stable")
        else:
            quality_indicators.append("⚠️ Écart WACC-Croissance faible : Sensibilité élevée")
        
        for indicator in quality_indicators:
            print(f"  {indicator}")
        
        print(f"\n🔍 RECOMMANDATIONS")
        print("-" * 30)
        
        recommendations = [
            "• Valider hypothèses avec comparables sectoriels",
            "• Effectuer due diligence sur drivers de valeur",
            "• Mettre à jour avec données de marché récentes"
        ]
        
        if self.results['terminal_percentage'] > 70:
            recommendations.append("• Réduire période projection ou ajuster hypothèses terminales")
        
        if self.results['avg_roic'] < self.assumptions['wacc']:
            recommendations.append("• Analyser leviers d'amélioration rentabilité")
        
        for rec in recommendations:
            print(f"  {rec}")
        
        print("\n" + "=" * 100)
        print("🚀 ANALYSE DCF UNIVERSELLE TERMINÉE")
        print(f"📊 Mode: {'COMPLET' if all([HAS_MATPLOTLIB, HAS_OPENPYXL]) else 'ADAPTATIF'} - Toutes validations effectuées")
        print("=" * 100)
    
    def run_universal_analysis(self):
        """Lance l'analyse DCF universelle complète"""
        try:
            self.display_banner()
            
            # Informations de base
            print("🏢 INFORMATIONS ENTREPRISE")
            print("-" * 30)
            self.company_name = input("Nom de l'entreprise: ").strip() or "Entreprise Cible"
            
            # Sélection industrie
            print("\n🏭 SÉLECTION DE L'INDUSTRIE")
            print("-" * 40)
            industries = list(INDUSTRY_BENCHMARKS.keys())
            for i, industry in enumerate(industries, 1):
                print(f"  {i}. {industry}")
            
            try:
                choice = int(input(f"\nChoisissez l'industrie (1-{len(industries)}) [7]: ") or "7")
                if 1 <= choice <= len(industries):
                    self.industry = industries[choice - 1]
                else:
                    self.industry = "General"
            except:
                self.industry = "General"
            
            print(f"✅ Configuration: {self.company_name} - {self.industry}")
            
            # Collecte des hypothèses
            if not self.get_detailed_assumptions():
                return
            
            # Calculs principaux
            self.calculate_advanced_projections()
            self.calculate_dcf_valuation_advanced()
            
            # Analyses avancées
            print("\n🔄 ANALYSES AVANCÉES EN COURS...")
            
            self.advanced_scenario_analysis()
            self.sensitivity_analysis_2d()
            
            # Monte Carlo (adaptatif)
            if HAS_NUMPY:
                run_monte_carlo = input("\n🎲 Lancer simulation Monte Carlo? (o/n) [o]: ").lower()
                if run_monte_carlo != 'n':
                    self.monte_carlo_simulation()
            else:
                print("\n⚠️ Monte Carlo non disponible (NumPy requis)")
            
            # Valorisation comparative
            self.comparative_valuation()
            
            # Affichage des résultats
            self.display_professional_results()
            
            # Visualisations (si disponibles)
            if HAS_MATPLOTLIB:
                create_viz = input("\n📈 Générer les visualisations? (o/n) [o]: ").lower()
                if create_viz != 'n':
                    self.visualizer.create_visualizations(self)
            else:
                print("\n⚠️ Visualisations non disponibles")
                print("Pour activer: pip install matplotlib seaborn")
            
            # Export (adaptatif)
            export_data = input("\n📊 Exporter les résultats? (o/n) [o]: ").lower()
            if export_data != 'n':
                self.exporter.export_results(self)
            
            print("\n🎉 ANALYSE DCF UNIVERSELLE TERMINÉE AVEC SUCCÈS!")
            print(f"📊 Mode: {'COMPLET' if all([HAS_MATPLOTLIB, HAS_OPENPYXL, HAS_TABULATE]) else 'ADAPTATIF'}")
            
            # Instructions d'amélioration
            if not all([HAS_MATPLOTLIB, HAS_OPENPYXL, HAS_TABULATE, HAS_NUMPY]):
                print("\n🚀 POUR ACTIVER TOUTES LES FONCTIONNALITÉS:")
                print("pip install matplotlib pandas numpy tabulate openpyxl seaborn")
            
        except KeyboardInterrupt:
            print("\n⚠️ Analyse interrompue par l'utilisateur")
        except Exception as e:
            print(f"\n❌ Erreur inattendue: {e}")
            import traceback
            traceback.print_exc()

# ==================== MAIN ====================

def main():
    """Fonction principale universelle"""
    print("🚀 Démarrage du Calculateur DCF Universel...")
    print("🔄 Détection automatique des modules disponibles...")
    print()
    
    calculator = UniversalDCFCalculator()
    calculator.run_universal_analysis()

if __name__ == "__main__":
    main()